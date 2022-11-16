#! /usr/bin/env python3.6
# -*- coding: utf-8 -*-
import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QTableWidgetItem
from PyQt5.QtCore import QTimer
import sqlite3
from openpyxl import Workbook
import glavU
from elN import El_Window
from spisan import MWspis



class GL_Window(QtWidgets.QMainWindow, glavU.Ui_MW):
    def __init__(self):  # для доступа к элементам таблицы в файле t1

        super().__init__()
        self.setupUi(self)  # для инициализации дизайна

        self.t = QTimer(self)
        self.t.timeout.connect(self.t_out)
    t1 = 0

    def t_out(self):
        self.l5.setText("")
        self.t.stop()

    def snowElt(self):
        self.tw = El_Window()
        self.tw.show()

    def sSpis(self):
        self.sp = MWspis()
        self.sp.show()

    def raschetAll(self, s):
        self.rrr1 = []
        self.lenT1 = ' '
        if s == 'd':
            self.s1 = 'ДВЕ'
            j = 5   # номер столбца с количеством примен. элементов в ДВЕ
            self.lE3_2.setText('dve')

        if s == 'b':
            self.s1 = 'БС'
            j = 6   # номер столбца с количеством примен. элементов в БС
            self.lE3_2.setText('bs')


        # создание таблицы
        self.tableWidget.setColumnCount(6)  # колич столбцов
        self.tableWidget.setColumnWidth(0, 120)  # ширина столбцов
        self.tableWidget.setColumnWidth(1, 120)
        self.tableWidget.setColumnWidth(2, 120)
        self.tableWidget.setColumnWidth(3, 120)
        self.tableWidget.setColumnWidth(4, 120)
        self.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)  # запрет на редактирование
        self.tableWidget.horizontalHeader().setStretchLastSection(True)  #
        s = "::section{Background-color:#aaffff}"  # выбор цвета заголовка
        self.tableWidget.horizontalHeader().setStyleSheet(s)
        # self.tableWidget.horizontalHeader().setVisible(False)
        self.tableWidget.setHorizontalHeaderLabels(["Номер", "Наименование", "Номинал", "Количество", "Изгот. "+self.s1, "Примечания"])
        self.tableWidget.setRowCount(0)
        self.tableWidget.verticalHeader().setVisible(False)  #

        # Создание таблицы расчета из имеющихся элементов
        try:
            conn = sqlite3.connect('elements.db')
            cur = conn.cursor()
            cur.execute("SELECT * FROM element")
            row1 = cur.fetchall()
            cur.close()

            # расчет количества изготовленных приборов из всех элементов
            for i in range(len(row1)):

                if row1[i][j] == 0:
                    pass

                else:
                    self.r1 = row1[i][4] // row1[i][j]
                    self.rrr1.append([row1[i][0], self.r1])

            self.t.start(1000)  # задержка надписи

        except:
            self.l5.setText('<font color= "red">Ошибка расчета!</font>')
        else:
            self.l5.setText('<font color= "green">Расчет произведен!</font>')

        self.rrr1.sort(key=lambda x: x[1])  # сортировка
        self.lenT1 = len(self.rrr1)     # количество строк в таблице
        self.tableWidget.setRowCount(len(self.rrr1))
        # заполнение таблицы
        for i in range(len(self.rrr1)):
            k = self.rrr1[i][0] - 1
            self.tableWidget.setItem(i, 0, QTableWidgetItem(str(row1[k][1])))
            self.tableWidget.setItem(i, 1, QTableWidgetItem(str(row1[k][2])))
            self.tableWidget.setItem(i, 2, QTableWidgetItem(str(row1[k][3])))
            self.tableWidget.setItem(i, 3, QTableWidgetItem(str(row1[k][4])))
            self.tableWidget.setItem(i, 4, QTableWidgetItem(str(self.rrr1[i][1])))
            self.tableWidget.setItem(i, 5, QTableWidgetItem(str(row1[k][7])))

        GL_Window.t1 = 1       # таблица 1-го типа

    #def scorBD(self):
        #pass

    def srscN(self):    # неиспользуемый блок
        #pass
        self.l5.setText('В разработке')
        self.tableWidget.setRowCount(0)

    def srscImD(self):
        s = 'd'     # запись в заголовок таблицы "БС"
        GL_Window.raschetAll(self, s)

    def srscImB(self):
        s = 'b'     # запись в заголовок таблицы "ДВЕ"
        GL_Window.raschetAll(self, s)

    # запись в файл
    def sZfile(self):
        if GL_Window.t1 == 1:
            try:
                wb = Workbook()
                # ws = wb.active
                ws = wb.create_sheet("DVE", 0)
                ws.freeze_panes = 'A2'  # закрепленная верхняя строка
                ws.column_dimensions['B'].width = 17
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 35
                ws['A1'] = 'Номер'
                ws['B1'] = 'Наимен'
                ws['C1'] = 'Номинал'
                ws['D1'] = 'Кол-во'
                ws['E1'] = 'Изготов. ' + self.s1
                ws['F1'] = 'Примеч'

                stolb = ['A', 'B', 'C', 'D', 'E', 'F']  # обозн столбцов
                k = 0  # столбец таблицы программы
                for i in range(2, self.lenT1 + 2):  # перебор строк листа эксель

                    for bukva in stolb:  # перебор столб листа эксель
                        tb = bukva + str(i)

                        ws[tb] = self.tableWidget.item(i - 2, k).text()
                        k += 1
                    k = 0

                f = self.lE3_2.text()
                if f == '':
                    self.l5.setText('<font color= "red">Bведите имя файла!</font>')
                else:
                    f = f + '.xlsx'
                    wb.save(f)
                    self.l5.setText('<font color= "green">Файл записан!</font>')
                    self.t.start(1000)
            except:
                self.l5.setText('<font color= "red">Рассчитайте таблицу!</font>')
            else:
                pass

        if GL_Window.t1 == 2:
            try:
                wb = Workbook()
                #ws = wb.active
                ws = wb.create_sheet("остаток", 0)
                ws.freeze_panes = 'A2'  # закрепленная верхняя строка
                ws.column_dimensions['B'].width = 17
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 35
                ws['A1'] = 'Номер'
                ws['B1'] = 'Наимен'
                ws['C1'] = 'Номинал'
                ws['D1'] = 'Кол-во'
                ws['E1'] = 'Для ДВЕ'
                ws['F1'] = 'Для БС'
                ws['G1'] = 'Всего'
                ws['H1'] = 'Остаток'
                ws['I1'] = 'Дефицит'
                ws['J1'] = 'Примечания'
                ws['D2'] = self.s2
                ws.merge_cells('D2:G2')

                stolb = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']  # обозн столбцов
                k = 0   # номер столбца

                for i in range(3, self.lenT1 + 2):  # перебор строк листа эксель

                    for bukva in stolb:     # перебор столб листа эксель
                        tb = bukva + str(i)
                        ws[tb] = self.tableWidget.item(i-3, k).text()
                        k += 1
                    k = 0


                f = self.lE3_2.text()
                if f == '':
                    self.l5.setText('<font color= "red">Bведите имя файла!</font>')
                else:
                    f = f + '.xlsx'
                    wb.save(f)
                    self.l5.setText('<font color= "green">Файл записан!</font>')
                    self.t.start(1000)
            except:
                self.l5.setText('<font color= "red">Рассчитайте таблицу!</font>')
            else:
                pass

    # расчет необходимого кол-ва элементов для приборов
    def spBA(self):
        # создание таблицы
        self.tableWidget.setColumnCount(10)  # колич столбцов
        self.tableWidget.setColumnWidth(0, 100)  # ширина столбцов
        self.tableWidget.setColumnWidth(1, 120)
        self.tableWidget.setColumnWidth(2, 120)
        self.tableWidget.setColumnWidth(3, 100)
        self.tableWidget.setColumnWidth(4, 80)
        self.tableWidget.setColumnWidth(5, 80)
        self.tableWidget.setColumnWidth(6, 80)
        self.tableWidget.setColumnWidth(7, 80)
        self.tableWidget.setColumnWidth(8, 80)
        self.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)  # запрет на редактирование
        self.tableWidget.horizontalHeader().setStretchLastSection(True)  #
        s = "::section{Background-color:#aaffff}"  # выбор цвета заголовка
        self.tableWidget.horizontalHeader().setStyleSheet(s)
        # self.tableWidget.horizontalHeader().setVisible(False)
        self.tableWidget.setHorizontalHeaderLabels(
             ["Номер", "Наименование", "Номинал", "Количество", "Для ДВЕ", "Для БС", "Всего", "Остаток", "Дефицит", "Примечания"])
        self.tableWidget.setRowCount(0)
        self.tableWidget.verticalHeader().setVisible(False)  #

        # Создание таблицы расчета из имеющихся элементов raschet1
        try:
            conn = sqlite3.connect('elements.db')
            cur = conn.cursor()
            cur.execute("SELECT * FROM element")
            row1 = cur.fetchall()
            cur.close()
        except:
            self.l5.setText('<font color= "red">Ошибка Базы данных!</font>')
        else:

            # расчет количества изготовленных приборов из всех элементов
            self.rdve = int(self.lE1.text())    # задание количества ДВЕ
            self.rbs = int(self.lE2.text())     # задание количества БС
            self.tableWidget.setRowCount(len(row1))

            #заполнение таблицы из БД
            for i in range(len(row1)):
                for j in range(4):
                    self.tableWidget.setItem(i, j, QTableWidgetItem(str(row1[i][j+1])))
                    self.tableWidget.setItem(i, 9, QTableWidgetItem(str(row1[i][7])))

            # расчет и заполнение таблицы
            for i in range(len(row1)):
                a1 = row1[i][5] * self.rdve
                a2 = row1[i][6] * self.rbs
                a3 = a1 + a2
                a4 = row1[i][4] - a3
                if a4 < 0:
                    a5 = abs(a4)
                else:
                    a5 = ''
                self.tableWidget.setItem(i, 4, QTableWidgetItem(str(a1)))
                self.tableWidget.setItem(i, 5, QTableWidgetItem(str(a2)))
                self.tableWidget.setItem(i, 6, QTableWidgetItem(str(a3)))
                self.tableWidget.setItem(i, 7, QTableWidgetItem(str(a4)))
                self.tableWidget.setItem(i, 8, QTableWidgetItem(str(a5)))

            self.s2 = ' Расчет произведен для '+str(self.rdve)+' ДВЕ и '+str(self.rbs)+' БС'
            self.l5.setText(self.s2)
            GL_Window.t1 = 2   # таблица 2-го типа
            self.lE3_2.setText('ostatok')

            self.lenT1 = len(row1)  # количество строк в таблице


def main():
    app = QtWidgets.QApplication(sys.argv)  # новый экземпляр QApplication
    w = GL_Window()

    w.show()

    sys.exit(app.exec_())
# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()



# See PyCharm help at https://www.jetbrains.com/help/pycharm/
